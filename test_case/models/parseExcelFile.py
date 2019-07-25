from openpyxl import load_workbook
from config.conf import dataPath

class ParseExcel(object):
    def __init__(self):
        self.wk = load_workbook(dataPath)
        self.excelFile = dataPath

    def getSheetByName(self,sheetName):
        """获取sheet对象"""
        sheet= self.wk[sheetName]
        return sheet

    def getRowNum(self,sheet):
        """获取有效数据的最大行号"""
        return sheet.max_row

    def getColsNum(self,sheet):
        """获取有效数据的最大列好"""
        return sheet.max_column

    def getRowValues(self,sheet,rowNum):
        """获取某一行的数据"""
        maxColsNum = self.getColsNum(sheet)
        rowValues = []
        for colsNum in range(1,maxColsNum + 1):
            value = sheet.cell(rowNum,colsNum).value
            if value is None:
                value = ' '
            rowValues.append(value)
        return tuple(rowValues)


    def getColumnValues(self,sheet,columNum):
        """获取某一列的数据"""
        maxRowNum = self.getRowNum(sheet)
        columnValues = []
        for rowNum in range(2,maxRowNum + 1):
            value = sheet.cell(rowNum,columNum).value
            if value is None:
                value = " "
            columnValues.append(value)
        return tuple(columnValues)


    def getValueOfCell(self,sheet,rowNum,columnNum):
        """获取某一个单元格的数据"""
        value = sheet.cell(rowNum,columnNum).value
        if value is None:
            value = " "
            return value

    def getAllValueOfSheet(self,sheet):
        """获取某一个sheet页的所有测试数据，返回一个元组组成的列表"""
        maxRowNum = self.getRowNum(sheet)
        columnNum = self.getColsNum(sheet)
        allValues = []
        for row in range (2,maxRowNum+1):
            rowValues = []
            for column in range(1,columnNum + 1):
                value = sheet.cell(row,column).value
                if value is None :
                    value= " "
                rowValues.append(value)
            allValues.append(tuple(rowValues))
        return allValues

if __name__ =="__main__":
    pef = ParseExcel()
    sheet = pef.getSheetByName("login")
    print(sheet)
    print(pef.getRowValues(sheet,2)[0])
    a = pef.getAllValueOfSheet(sheet)
    print(a)